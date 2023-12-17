import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import platform
import socket
import psutil
import winreg
import sqlite3
import appdirs
import re
import json
import base64
import win32crypt
from Cryptodome.Cipher import AES
import shutil

# Function to run a command and capture the output
def run_command(command):
    try:
        result = subprocess.check_output(command, shell=True, text=True)
        return result.strip()
    except subprocess.CalledProcessError:
        return None


# Function to create a styled Excel sheet
def create_styled_excel():
    wb = Workbook()
    ws = wb.active

    # Apply formatting to the header row
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(size=14, bold=True)

    # Apply formatting to the data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(size=12)

    return wb, ws


# Function to check if Windows Defender is enabled
def is_windows_defender_enabled():
    try:
        result = subprocess.check_output(
            [
                "powershell",
                "Get-MpPreference | Select-Object -ExpandProperty RealTimeScanEnabled",
            ]
        )
        return result.strip().decode("utf-8") == "True"
    except subprocess.CalledProcessError:
        return None


# Function to get environment variables
def get_environment_variables():
    environment_variables = os.environ
    return environment_variables


# Function to get critical events from Windows Event Log
def get_critical_events():
    critical_events = []
    with subprocess.Popen(
        [
            "powershell",
            'Get-WinEvent -FilterHashTable @{LogName="System"; Level=1,2,3} | Select-Object TimeCreated,Id,Message',
        ],
        stdout=subprocess.PIPE,
        text=True,
    ) as proc:
        for line in proc.stdout:
            critical_events.append(line.strip())
    return critical_events


# Function to collect system information
def collect_system_information():
    output_file = os.path.join("AllOutput", "system_information_output.txt")

    with open(output_file, "w") as text_file:
        text_file.write("System Information:\n")

        # Hostname
        hostname = socket.gethostname()
        text_file.write(f"Hostname: {hostname}\n")

        # System platform
        system_platform = platform.system()
        text_file.write(f"System Platform: {system_platform}\n")

        # Processor information
        processor_info = platform.processor()
        text_file.write(f"Processor: {processor_info}\n")

        # Memory information
        memory_info = psutil.virtual_memory()
        text_file.write(f"Total Memory: {memory_info.total / (1024 ** 3):.2f} GB\n")
        text_file.write(
            f"Available Memory: {memory_info.available / (1024 ** 3):.2f} GB\n"
        )

        # Disk information
        disk_info = psutil.disk_usage("/")
        text_file.write(f"Total Disk Space: {disk_info.total / (1024 ** 3):.2f} GB\n")
        text_file.write(f"Used Disk Space: {disk_info.used / (1024 ** 3):.2f} GB\n")

        # Network information
        network_info = psutil.net_if_addrs()
        text_file.write("Network Information:\n")
        for interface, addresses in network_info.items():
            text_file.write(f"Interface: {interface}\n")
            for address in addresses:
                text_file.write(f"  - Address Family: {address.family.name}\n")
                text_file.write(f"    Address: {address.address}\n")

        # Operating system version
        os_version = platform.version()
        text_file.write(f"Operating System Version: {os_version}\n")

        # Windows installation date (requires elevated permissions)
        try:
            with winreg.OpenKey(
                winreg.HKEY_LOCAL_MACHINE,
                r"SOFTWARE\Microsoft\Windows NT\CurrentVersion",
            ) as key:
                install_date = winreg.QueryValueEx(key, "InstallDate")[0]
                text_file.write(f"Windows Installation Date: {install_date}\n")
        except Exception as e:
            text_file.write(f"Error retrieving Windows installation date: {str(e)}\n")

        text_file.write("System Information collected successfully.")


# Function to collect environment information
def collect_environment_information(output_file):
    print("Collecting Environment Information...")
    environment_variables = get_environment_variables()
    output_text = "Environment Information:\n"

    for key, value in environment_variables.items():
        output_text += f"{key}: {value}\n"

    # Specify the output directory
    output_directory = "AllOutput"
    os.makedirs(output_directory, exist_ok=True)

    # Save the collected information to the specified output file
    with open(os.path.join(output_directory, output_file), "w") as text_file:
        text_file.write(output_text)


# Function to collect security events
def collect_security_events(output_file):
    print("Collecting Security Events...")

    # Get critical events
    critical_events = get_critical_events()

    # Create a DataFrame for the collected information
    data = {"Security Events": critical_events}

    # Create a DataFrame from the collected data
    output_df = pd.DataFrame(data)

    # Specify the output directory
    output_directory = "AllOutput"
    os.makedirs(output_directory, exist_ok=True)

    # Save the DataFrame to an Excel file
    excel_output_path = os.path.join(output_directory, output_file)
    output_df.to_excel(excel_output_path, index=False)


# ... (other task collection functions)

# -------------------------------------------

# Function to list connected USB devices and get detailed information about running processes
def list_usb_devices():
    usb_devices = []
    for device in psutil.disk_partitions():
        if "removable" in device.opts or "cdrom" in device.opts:
            usb_devices.append(device.device)
    return usb_devices


# Function to get detailed information about running processes
def get_running_processes():
    running_processes = []
    for process in psutil.process_iter(["pid", "name", "username", "memory_info"]):
        running_processes.append(
            {
                "PID": process.info["pid"],
                "Name": process.info["name"],
                "Username": process.info["username"],
                "Memory Usage (MB)": process.info["memory_info"].rss / (1024**2),
            }
        )
    return running_processes


# Function to get details about active network connections
def get_active_network_connections():
    active_network_connections = []
    for connection in psutil.net_connections(kind="inet"):
        local_address = f"{connection.laddr.ip if connection.laddr else 'N/A'}:{connection.laddr.port if connection.laddr else 'N/A'}"
        remote_address = f"{connection.raddr.ip if connection.raddr else 'N/A'}:{connection.raddr.port if connection.raddr else 'N/A'}"
        active_network_connections.append(
            {
                "Local Address": local_address,
                "Remote Address": remote_address,
                "Status": connection.status,
                "PID": connection.pid if connection.pid else "N/A",
            }
        )
    return active_network_connections


# Function to get information about active user sessions
def get_active_user_sessions():
    active_user_sessions = []
    for session in psutil.users():
        active_user_sessions.append(
            {
                "Username": session.name,
                "Terminal": session.terminal,
                "Host": session.host,
                "Started": session.started,
            }
        )
    return active_user_sessions


# Function to get information from Windows Registry
def get_registry_info(key, subkey):
    registry_info = []
    try:
        hkey = getattr(winreg, key)
        with winreg.OpenKey(hkey, subkey) as reg_key:
            i = 0
            while True:
                name, value, _ = winreg.EnumValue(reg_key, i)
                registry_info.append({"Name": name, "Value": value})
                i += 1
    except Exception as e:
        pass
    return registry_info


# Function to get information about installed Windows updates
def get_installed_updates():
    try:
        result = subprocess.check_output(["wmic", "qfe", "list", "full"])
        return result.strip().decode("utf-8")
    except subprocess.CalledProcessError:
        return None


def collect_usb_devices_and_processes(output_file):
    print("Collecting USB Devices and Processes...")

    # List connected USB devices
    usb_devices = list_usb_devices()

    # Get detailed information about running processes
    running_processes = get_running_processes()

    # Specify the output directory
    output_directory = "AllOutput"
    os.makedirs(output_directory, exist_ok=True)

    # Save USB devices to the specified output file
    usb_devices_output_file = os.path.join(
        output_directory, "usb_devices_output." + output_file.split(".")[-1]
    )
    with open(usb_devices_output_file, "w") as usb_file:
        usb_file.write("Connected USB Devices:\n")
        for usb_device in usb_devices:
            usb_file.write(f"{usb_device}\n")

    # Save running processes to the specified output file
    processes_output_file = os.path.join(
        output_directory, "running_processes_output." + output_file.split(".")[-1]
    )
    with open(processes_output_file, "w") as process_file:
        process_file.write("Running Processes:\n")
        for process in running_processes:
            process_file.write(
                f"PID: {process['PID']}, Name: {process['Name']}, Username: {process['Username']}, Memory Usage (MB): {process['Memory Usage (MB)']:.2f}\n"
            )

    print("USB Devices and Processes collected successfully.")


# Function to get information about active user sessions and installed updates
def collect_user_sessions_and_updates(output_file):
    print("Collecting User Sessions and Updates...")

    # Get information about active user sessions
    active_user_sessions = get_active_user_sessions()

    # Get information about installed Windows updates
    installed_updates = get_installed_updates()

    # Specify the output directory
    output_directory = "AllOutput"
    os.makedirs(output_directory, exist_ok=True)

    # Save active user sessions to the specified output file
    user_sessions_output_file = os.path.join(
        output_directory, "active_user_sessions_output." + output_file.split(".")[-1]
    )
    with open(user_sessions_output_file, "w") as user_session_file:
        user_session_file.write("Active User Sessions:\n")
        for session in active_user_sessions:
            user_session_file.write(
                f"Username: {session['Username']}, Terminal: {session['Terminal']}, Host: {session['Host']}, Started: {session['Started']}\n"
            )

    # Save installed updates to the specified output file
    updates_output_file = os.path.join(
        output_directory, "installed_updates_output." + output_file.split(".")[-1]
    )
    with open(updates_output_file, "w") as update_file:
        update_file.write("Installed Windows Updates:\n")
        if installed_updates:
            update_file.write(installed_updates)

    print("User Sessions and Updates collected successfully.")


# Function to get information from Windows Registry and active network connections
def collect_registry_and_network_information(output_file):
    print("Collecting Registry and Network Information...")

    # Get information from Windows Registry
    registry_info = get_registry_info(
        "HKEY_LOCAL_MACHINE", r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"
    )

    # Get details about active network connections
    active_network_connections = get_active_network_connections()

    # Specify the output directory
    output_directory = "AllOutput"
    os.makedirs(output_directory, exist_ok=True)

    # Save Registry information to the specified output file
    registry_output_file = os.path.join(
        output_directory, "registry_information_output." + output_file.split(".")[-1]
    )
    with open(registry_output_file, "w") as registry_file:
        registry_file.write("Windows Registry Information:\n")
        for reg_entry in registry_info:
            registry_file.write(
                f"Registry Entry - {reg_entry['Name']}: {reg_entry['Value']}\n"
            )

    # Save active network connections to the specified output file
    network_output_file = os.path.join(
        output_directory,
        "active_network_connections_output." + output_file.split(".")[-1],
    )
    with open(network_output_file, "w") as network_file:
        network_file.write("Active Network Connections:\n")
        for connection in active_network_connections:
            network_file.write(
                f"Local Address: {connection['Local Address']}, Remote Address: {connection['Remote Address']}, Status: {connection['Status']}, PID: {connection['PID']}\n"
            )


# Function to list scheduled tasks
def save_scheduled_tasks_to_excel(output_directory):
    try:
        # Use the schtasks command to list scheduled tasks
        result = subprocess.check_output(
            ["schtasks", "/query", "/fo", "csv"], text=True
        )

        # Split the CSV lines into a list
        lines = result.strip().split("\n")

        # Extract header and task data
        header = lines[0].split(",")
        task_data = [line.split(",") for line in lines[1:]]

        # Create a list of dictionaries representing tasks
        tasks = [dict(zip(header, task)) for task in task_data]

        # Create a DataFrame from the list of tasks
        tasks_df = pd.DataFrame(tasks)

        # Create the output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)

        # Specify the output file path within the output directory
        excel_output_file = os.path.join(
            output_directory, "scheduled_tasks_output.xlsx"
        )

        # Save the DataFrame to an Excel file
        tasks_df.to_excel(excel_output_file, index=False)

        print(f"Scheduled tasks saved to {excel_output_file}")

    except subprocess.CalledProcessError as e:
        print(f"Error listing scheduled tasks: {e}")


# Function to run network analysis


def run_commands_and_save_to_txt(commands, output_directory, output_filename):
    # Create the output directory if it doesn't exist
    os.makedirs(output_directory, exist_ok=True)

    # Specify the output file path within the output directory
    txt_file = os.path.join(output_directory, output_filename)

    try:
        with open(txt_file, "w") as file:
            for command in commands:
                # Run the command
                output = run_command(command)

                # Save the results to the text file
                file.write("\n" + "=" * 20 + "\n")
                file.write(f"Results for command: {command}\n")
                if output:
                    file.write(output)
                else:
                    file.write("No output\n")

        print(f"Results saved to {txt_file}")

    except Exception as e:
        print(f"An error occurred: {e}")


def run_command(command):
    try:
        result = subprocess.run(
            command,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        return result.stdout
    except Exception as e:
        return f"Error executing command: {str(e)}"


# Example usage:
commands_to_run = [
    "wevtutil qe System /rd:true /f:text /c:5",
    "wevtutil qe Security /rd:true /f:text /c:5",
    "wevtutil qe Application /rd:true /f:text /c:5",
    "netstat -ano",
    "netstat -n",
    "netsh advfirewall firewall show rule name=all",
    "ipconfig /displaydns",
    "arp -a",
]


# Function to perform Chrome browser analysis
def chrome_browser_analysis():
    # Your Chrome browser analysis code here
    pass


# Function to collect Chrome search history
def extract_and_save_search_history():
    try:
        # Get the user data directory for Chrome
        data_dir = appdirs.user_data_dir()

        # Construct the Chrome history database path
        history_db_path = os.path.join(
            data_dir, "Google", "Chrome", "User Data", "Default", "History"
        )

        # Connect to the Chrome history database
        conn = sqlite3.connect(history_db_path)
        cursor = conn.cursor()

        # Retrieve search history from the database
        cursor.execute(
            "SELECT url, title, datetime((last_visit_time/1000000)-11644473600, 'unixepoch', 'localtime') AS last_visit_time FROM urls"
        )
        search_history = cursor.fetchall()

        # Create a pandas DataFrame from the retrieved search history
        df = pd.DataFrame(search_history, columns=["url", "title", "Timestamp"])

        # Specify the output directory
        output_directory = "AllOutput"

        # Create the output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)

        # Specify the output file path within the output directory
        excel_output_file = os.path.join(output_directory, "search_history.xlsx")

        # Save the search history DataFrame to an Excel file
        df.to_excel(excel_output_file, index=False)

        print(f"Search history saved to {excel_output_file}")

    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Close the database connection in the 'finally' block to ensure it happens even if an error occurs.
        if conn:
            conn.close()


# ------------------------------
# GLOBAL CONSTANT
CHROME_PATH_LOCAL_STATE = os.path.normpath(
    r"%s\AppData\Local\Google\Chrome\User Data\Local State"
    % (os.environ["USERPROFILE"])
)
CHROME_PATH = os.path.normpath(
    r"%s\AppData\Local\Google\Chrome\User Data" % (os.environ["USERPROFILE"])
)


def get_secret_key():
    try:
        with open(CHROME_PATH_LOCAL_STATE, "r", encoding="utf-8") as f:
            local_state = f.read()
            local_state = json.loads(local_state)
        secret_key = base64.b64decode(local_state["os_crypt"]["encrypted_key"])
        secret_key = secret_key[5:]
        secret_key = win32crypt.CryptUnprotectData(secret_key, None, None, None, 0)[1]
        return secret_key
    except Exception as e:
        print("%s" % str(e))
        print("[ERR] Chrome secretkey cannot be found")
        return None


def decrypt_payload(cipher, payload):
    return cipher.decrypt(payload)


def generate_cipher(aes_key, iv):
    return AES.new(aes_key, AES.MODE_GCM, iv)


def decrypt_password(ciphertext, secret_key):
    try:
        initialisation_vector = ciphertext[3:15]
        encrypted_password = ciphertext[15:-16]
        cipher = generate_cipher(secret_key, initialisation_vector)
        decrypted_pass = decrypt_payload(cipher, encrypted_password)
        decrypted_pass = decrypted_pass.decode()
        return decrypted_pass
    except Exception as e:
        print("%s" % str(e))
        print(
            "[ERR] Unable to decrypt, Chrome version <80 not supported. Please check."
        )
        return ""


def get_db_connection(chrome_path_login_db):
    try:
        shutil.copy2(chrome_path_login_db, "Loginvault.db")
        return sqlite3.connect("Loginvault.db")
    except Exception as e:
        print("%s" % str(e))
        print("[ERR] Chrome database cannot be found")
        return None


def extract_and_save_passwords():
    try:
        # Create a list to store the extracted passwords
        passwords_data = []

        # (1) Get secret key
        secret_key = get_secret_key()

        # Search user profile or default folder (this is where the encrypted login password is stored)
        folders = [
            element
            for element in os.listdir(CHROME_PATH)
            if re.search("^Profile*|^Default$", element) != None
        ]
        for folder in folders:
            # (2) Get ciphertext from sqlite database
            chrome_path_login_db = os.path.normpath(
                r"%s\%s\Login Data" % (CHROME_PATH, folder)
            )
            conn = get_db_connection(chrome_path_login_db)
            if secret_key and conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT action_url, username_value, password_value FROM logins"
                )
                for index, login in enumerate(cursor.fetchall()):
                    url = login[0]
                    username = login[1]
                    ciphertext = login[2]
                    if url != "" and username != "" and ciphertext != "":
                        decrypted_password = decrypt_password(ciphertext, secret_key)

                        passwords_data.append(
                            {
                                "Sequence": index,
                                "URL": url,
                                "Username": username,
                                "Password": decrypted_password,
                            }
                        )

                        # print("Sequence: %d" % index)
                        # print("URL: %s\nUser Name: %s\nPassword: %s\n" % (url, username, decrypted_password))
                        # print("*" * 50)

                # Close database connection
                cursor.close()
                conn.close()
                # Delete temp login db
                os.remove("Loginvault.db")

        # Convert the list of dictionaries to a DataFrame
        passwords_df = pd.DataFrame(passwords_data)

        # Specify the output directory
        output_directory = "AllOutput"

        # Create the output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)

        # Specify the output file path within the output directory
        excel_output_file = os.path.join(output_directory, "decrypted_password.xlsx")

        # Save the DataFrame to an Excel file
        passwords_df.to_excel(excel_output_file, index=False)

        print(f"Results saved to {excel_output_file}")

    except Exception as e:
        print("[ERR] %s" % str(e))

# Function to create the "AllOutput" folder if it doesn't exist
def create_output_folder():
    output_folder = "AllOutput"
    os.makedirs(output_folder, exist_ok=True)

def main():
    
    create_output_folder()
    while True:
        print("\nChoose a task to run:")
        print("1. Collect System Information")
        print("2. Collect Environment Information")
        print("3. Collect Security Events")
        print("4. USB Devices and Processes")
        print("5. User Sessions and Updates")
        print("6. Registry and Network Information")
        print("7. Scheduled Tasks")
        print("8. Network Analysis")
        print("9. Chrome Browser Passwords Analysis")
        print("10. Chrome Search History")
        print("A. Run All Tasks")
        print("0. Exit")

        choice = input("Enter the task number (0-10 or A): ")

        if choice == "1":
            output_file = "system_information_output.txt"
            collect_system_information()
            print(f"System Information saved to {output_file}")
        elif choice == "2":
            output_file = "environment_information_output.txt"
            collect_environment_information(output_file)
            print(f"Environment Information saved to {output_file}")
        elif choice == "3":
            output_file = "security_events_output.xlsx"
            collect_security_events("security_events_output.xlsx")
            print(f"Security Events saved to {output_file}")
        elif choice == "4":
            output_file = "USB_Devices_and_Processes_output.txt"
            collect_usb_devices_and_processes(output_file)

        elif choice == "5":
            output_file = "User_Sessions_and_Updates_output.txt"
            collect_user_sessions_and_updates(output_file)

        elif choice == "6":
            output_file = "Registry_and_Network_Information_output.txt"
            collect_registry_and_network_information(output_file)
        elif choice == "7":
            output_directory = "AllOutput"
            save_scheduled_tasks_to_excel(output_directory)
        elif choice == "8":
            output_directory = "AllOutput"
            output_filename = "network_analysis_results.txt"
            run_commands_and_save_to_txt(
                commands_to_run, output_directory, output_filename
            )
        elif choice == "9":
            extract_and_save_passwords()
        elif choice == "10":
            extract_and_save_search_history()

        elif choice.lower() == "a":
            output_directory = "AllOutput"
            # Task 1
            output_file = "system_information_output.txt"
            collect_system_information()
            print(f"System Information saved to {output_file}")

            # Task 2
            output_file = "environment_information_output.txt"
            collect_environment_information(output_file)
            print(f"Environment Information saved to {output_file}")

            # Task 3
            output_file = "security_events_output.xlsx"
            collect_security_events(output_file)
            print(f"Security Events saved to {output_file}")

            # Task 4
            output_file = "USB_Devices_and_Processes_output.txt"
            collect_usb_devices_and_processes(output_file)

            # Task 5
            output_file = "User_Sessions_and_Updates_output.txt"
            collect_user_sessions_and_updates(output_file)

            # Task 6
            output_file = "Registry_and_Network_Information_output.txt"
            collect_registry_and_network_information(output_file)

            # Task 7
            save_scheduled_tasks_to_excel(output_directory)

            # Task 8
            output_filename = "network_analysis_results.txt"
            run_commands_and_save_to_txt(
                commands_to_run, output_directory, output_filename
            )

            # Task 9
            extract_and_save_passwords()
            # Task 10
            extract_and_save_search_history()
        elif choice == "0":
            print("Exiting the script.")
            break
        else:
            print("Invalid choice. Please enter a valid option.")


if __name__ == "__main__":
    main()
